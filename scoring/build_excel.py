"""
build_excel.py — Job application tracker exporter.

Reads storage/application_history.json (the ledger written by src/history.mjs)
and emits an Excel workbook with:
  - One row per generated application (sorted newest-first).
  - Clickable links to the posting URL and to the local CV/CL PDFs.
  - A status column driven by a data-validation dropdown (Applied / Screening /
    Interviewing / Offer / Rejected / Withdrawn).
  - Column-width auto-fit and a frozen header row.

No network, no telemetry. Pure openpyxl.

Usage:
    python3 scoring/build_excel.py [--history path] [--out path]
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


HERE = Path(__file__).resolve().parent
ROOT = HERE.parent
DEFAULT_HISTORY = ROOT / "storage" / "application_history.json"
DEFAULT_OUT = ROOT / "output" / "applications.xlsx"

STATUS_OPTIONS = [
    "Applied", "Screening", "Interviewing", "Offer", "Rejected", "Withdrawn",
]

HEADERS = [
    "Generated", "Company", "Role", "Archetype", "Status",
    "Job URL", "CV (PDF)", "Cover Letter (PDF)", "Slug",
]


def _load_entries(path: Path) -> list[dict]:
    if not path.exists():
        return []
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return []
    entries = data.get("entries") if isinstance(data, dict) else data
    return entries if isinstance(entries, list) else []


def _fmt_date(iso: str) -> str:
    # Keep the human-readable slice (YYYY-MM-DD HH:MM); history stores full ISO.
    return (iso or "").replace("T", " ")[:16]


def build(history_path: Path, out_path: Path) -> Path:
    entries = _load_entries(history_path)
    # Newest first.
    entries.sort(key=lambda e: e.get("generatedAt", ""), reverse=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"

    # Header row.
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")
    for col, name in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=col, value=name)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="left", vertical="center")
    ws.freeze_panes = "A2"

    link_font = Font(color="2563EB", underline="single")

    def _link_cell(row: int, col: int, target: str, label: str):
        if not target:
            ws.cell(row=row, column=col, value="")
            return
        c = ws.cell(row=row, column=col, value=label or target)
        c.hyperlink = target
        c.font = link_font

    for i, e in enumerate(entries, start=2):
        ws.cell(row=i, column=1, value=_fmt_date(e.get("generatedAt", "")))
        ws.cell(row=i, column=2, value=e.get("company", ""))
        ws.cell(row=i, column=3, value=e.get("role", ""))
        ws.cell(row=i, column=4, value=e.get("archetype", ""))
        ws.cell(row=i, column=5, value="Applied")  # default; user edits via dropdown
        _link_cell(i, 6, e.get("url", ""), e.get("url", ""))
        _link_cell(i, 7, e.get("cvPath", ""), Path(e.get("cvPath", "")).name if e.get("cvPath") else "")
        _link_cell(i, 8, e.get("clPath", ""), Path(e.get("clPath", "")).name if e.get("clPath") else "")
        ws.cell(row=i, column=9, value=e.get("slug", ""))

    # Status dropdown on column E.
    last_row = max(len(entries) + 1, 2)
    dv = DataValidation(
        type="list",
        formula1='"{}"'.format(",".join(STATUS_OPTIONS)),
        allow_blank=True,
    )
    dv.add(f"E2:E{last_row}")
    ws.add_data_validation(dv)

    # Column widths — generous but capped.
    widths = {1: 18, 2: 26, 3: 34, 4: 20, 5: 14, 6: 50, 7: 32, 8: 32, 9: 28}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def main() -> int:
    ap = argparse.ArgumentParser(description="Export application history to Excel.")
    ap.add_argument("--history", type=Path, default=DEFAULT_HISTORY)
    ap.add_argument("--out", type=Path, default=DEFAULT_OUT)
    args = ap.parse_args()

    path = build(args.history, args.out)
    entries = _load_entries(args.history)
    print(f"Wrote {path} ({len(entries)} application(s)).")
    return 0


if __name__ == "__main__":
    sys.exit(main())
